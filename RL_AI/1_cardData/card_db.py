
from __future__ import annotations

from dataclasses import dataclass, field, asdict
from enum import IntEnum
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional
import re

try:
    import openpyxl
except ImportError as exc:  # pragma: no cover
    raise ImportError("card_db.py requires openpyxl. Install with `pip install openpyxl`.") from exc


class Role(IntEnum):
    LEADER = 0
    BISHOP = 1
    KNIGHT = 2
    ROOK = 3
    PAWN = 4


class Condition(IntEnum):
    ALWAYS = 0
    TURN_START = 1
    TURN_END = 2
    ON_DESTROY = 3
    ON_BASIC_MOVE = 4


ROLE_NAME_KO = {
    Role.LEADER: "군주",
    Role.BISHOP: "비숍",
    Role.KNIGHT: "나이트",
    Role.ROOK: "룩",
    Role.PAWN: "폰",
}

CONDITION_NAME_KO = {
    Condition.ALWAYS: "상시",
    Condition.TURN_START: "턴 시작",
    Condition.TURN_END: "턴 종료",
    Condition.ON_DESTROY: "파괴 시",
    Condition.ON_BASIC_MOVE: "기본 이동 시",
}

MODULE_DIR = Path(__file__).resolve().parent


def _resolve_module_relative(path: str | Path) -> Path:
    p = Path(path)
    if p.is_absolute():
        return p
    return (MODULE_DIR / p).resolve()



def _safe_int(value: Any) -> int:
    if value is None:
        return 0
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, (int, float)):
        return int(value)
    text = str(value).strip()
    if not text:
        return 0
    if text.startswith("0x"):
        return int(text, 16)
    return int(float(text))


def _safe_str(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def classify_target_schema(text: str) -> str:
    """
    액티브/패시브 텍스트에서 기본 대상 유형을 추정한다.
    완전한 룰 엔진은 아니고, legal action 생성 보조용 메타정보다.
    """
    text = _safe_str(text)
    if not text:
        return "none"

    if "최대 두 마리" in text or "최대 2" in text:
        return "enemy_multi_2"
    if "모든 적" in text or "같은 행의 적" in text:
        return "enemy_row"
    if "아군 군주를 선택" in text or "군주를 선택" in text:
        return "ally_leader"
    if "군주가 아닌 아군 기물" in text:
        return "ally_nonleader"
    if "아군 기물을 선택" in text:
        return "ally_unit"
    if "적 기물을 선택" in text or "공격 가능한 적 기물" in text:
        return "enemy_unit"
    if "원하는 칸" in text or "위치로 이동" in text or "옮깁니다" in text:
        return "board_cell"
    return "none"


def extract_effect_flags(text: str) -> List[str]:
    """
    텍스트에 등장하는 대표 효과 패턴을 멀티 라벨처럼 뽑아낸다.
    추후 rules/effects.py에서 정교한 파서를 붙이기 전까지의 임시 메타다.
    """
    text = _safe_str(text)
    flags: List[str] = []
    patterns = [
        ("auto_attack", r"기동범위 내 적.*공격"),
        ("draw", r"카드를 한 장 뽑"),
        ("gain_mana", r"마나를 \d+ 얻"),
        ("heal", r"회복"),
        ("full_heal", r"체력을 전부 회복"),
        ("shield", r"보호막"),
        ("buff_attack", r"공격력을 \d+ 올"),
        ("buff_move", r"더 이동할 수 있"),
        ("disable_move", r"움직일 수 없"),
        ("disable_attack", r"공격할 수 없"),
        ("swap", r"위치를 바꿉"),
        ("force_attack", r"공격 가능한 .* 공격합니다"),
        ("move_after_kill", r"파괴되었다면 .*위치로 이동"),
        ("promotion", r"프로모션|승격"),
        ("same_row_combo", r"군주와 같은 행"),
        ("sacrifice_ally", r"아군 기물을 하나 선택해 리타이어"),
        ("aoe", r"같은 행의 적 .* 피해"),
    ]
    for name, pattern in patterns:
        if re.search(pattern, text):
            flags.append(name)
    return flags


@dataclass(frozen=True)
class CardDefinition:
    card_id: str
    name: str
    world: int
    role: Role
    attack: int
    life: int
    summon_cost: int
    ability_cost: int
    condition: Condition
    passive_name: str
    passive_content: str
    active_name: str
    active_content: str
    passive_target_schema: str = "none"
    active_target_schema: str = "none"
    passive_flags: List[str] = field(default_factory=list)
    active_flags: List[str] = field(default_factory=list)

    @property
    def role_name(self) -> str:
        return ROLE_NAME_KO[self.role]

    @property
    def condition_name(self) -> str:
        return CONDITION_NAME_KO[self.condition]

    def to_dict(self) -> Dict[str, Any]:
        data = asdict(self)
        data["role"] = int(self.role)
        data["condition"] = int(self.condition)
        data["role_name"] = self.role_name
        data["condition_name"] = self.condition_name
        return data


def _normalize_row(row: Dict[str, Any]) -> CardDefinition:
    passive_content = _safe_str(row.get("PassiveContent"))
    active_content = _safe_str(row.get("ActiveContent"))

    return CardDefinition(
        card_id=_safe_str(row.get("CardID")),
        name=_safe_str(row.get("Name")),
        world=_safe_int(row.get("World")),
        role=Role(_safe_int(row.get("Role"))),
        attack=_safe_int(row.get("Attack")),
        life=_safe_int(row.get("Life")),
        summon_cost=_safe_int(row.get("SummonCost")),
        ability_cost=_safe_int(row.get("AbilityCost")),
        condition=Condition(_safe_int(row.get("Condition"))),
        passive_name=_safe_str(row.get("PassiveName")),
        passive_content=passive_content,
        active_name=_safe_str(row.get("ActiveName")),
        active_content=active_content,
        passive_target_schema=classify_target_schema(passive_content),
        active_target_schema=classify_target_schema(active_content),
        passive_flags=extract_effect_flags(passive_content),
        active_flags=extract_effect_flags(active_content),
    )


def load_card_db(xlsx_path: str | Path) -> Dict[str, CardDefinition]:
    """
    Cards.xlsx를 읽어서 {card_id: CardDefinition} 형태로 반환한다.
    시트 첫 4행의 설명문은 건너뛰고, 5행을 헤더로 사용한다.
    """
    xlsx_path = _resolve_module_relative(xlsx_path)
    workbook = openpyxl.load_workbook(xlsx_path, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]

    rows = list(sheet.iter_rows(values_only=True))
    if len(rows) < 6:
        raise ValueError("Cards.xlsx 형식이 예상과 다릅니다. 최소 6행 이상이 필요합니다.")

    header_row = rows[4]
    headers = [_safe_str(cell) for cell in header_row]

    required_headers = [
        "CardID", "Name", "World", "Role", "Attack", "Life",
        "SummonCost", "AbilityCost", "Condition",
        "PassiveName", "PassiveContent", "ActiveName", "ActiveContent",
    ]
    missing = [h for h in required_headers if h not in headers]
    if missing:
        raise ValueError(f"Cards.xlsx 헤더가 누락되었습니다: {missing}")

    db: Dict[str, CardDefinition] = {}
    for raw in rows[5:]:
        if raw is None or raw[0] is None:
            continue
        row = dict(zip(headers, raw[:len(headers)]))
        card = _normalize_row(row)
        db[card.card_id] = card

    return db


def load_card_list(xlsx_path: str | Path) -> List[CardDefinition]:
    return list(load_card_db(xlsx_path).values())


def group_cards_by_world(card_db: Dict[str, CardDefinition]) -> Dict[int, List[CardDefinition]]:
    grouped: Dict[int, List[CardDefinition]] = {}
    for card in card_db.values():
        grouped.setdefault(card.world, []).append(card)
    for world, cards in grouped.items():
        cards.sort(key=lambda c: (int(c.role), c.card_id))
    return grouped


def build_default_deck(card_db: Dict[str, CardDefinition], world: int) -> List[str]:
    """
    프로토타입 룰 기준:
    리더 1, 룩 1, 나이트 1, 비숍 1, 폰 3장(동일 카드)
    를 만족하는 기본 덱 카드 ID 리스트를 만든다.
    """
    world_cards = [c for c in card_db.values() if c.world == world]
    by_role: Dict[Role, List[CardDefinition]] = {}
    for card in world_cards:
        by_role.setdefault(card.role, []).append(card)

    required_roles = [Role.LEADER, Role.ROOK, Role.KNIGHT, Role.BISHOP, Role.PAWN]
    for role in required_roles:
        if role not in by_role or not by_role[role]:
            raise ValueError(f"world={world} 에 role={role.name} 카드가 없습니다.")

    leader = by_role[Role.LEADER][0]
    rook = by_role[Role.ROOK][0]
    knight = by_role[Role.KNIGHT][0]
    bishop = by_role[Role.BISHOP][0]
    pawn = by_role[Role.PAWN][0]
    return [leader.card_id, rook.card_id, knight.card_id, bishop.card_id, pawn.card_id, pawn.card_id, pawn.card_id]


def export_card_db_as_python_source(
    xlsx_path: str | Path,
    output_path: str | Path,
    variable_name: str = "CARD_DB_STATIC",
) -> Path:
    """
    Cards.xlsx를 읽어, 정적 파이썬 dict 코드로도 내보낸다.
    카드 수정이 잦지 않을 때 빠른 로딩용으로 쓸 수 있다.
    """
    output_path = _resolve_module_relative(output_path)
    card_db = load_card_db(xlsx_path)
    records = {cid: card.to_dict() for cid, card in sorted(card_db.items())}

    import pprint
    body = pprint.pformat(records, width=120, sort_dicts=True)
    source = (
        "# Auto-generated from Cards.xlsx\n"
        "# Do not edit manually unless you know what you are doing.\n\n"
        f"{variable_name} = {body}\n"
    )
    output_path.write_text(source, encoding="utf-8")
    return output_path


if __name__ == "__main__":
    # 예시 실행:
    # python card_db.py ./Cards.xlsx
    import sys
    xlsx = _resolve_module_relative(sys.argv[1]) if len(sys.argv) > 1 else _resolve_module_relative("Cards.xlsx")
    db = load_card_db(xlsx)
    print(f"loaded {len(db)} cards from {xlsx}")
    grouped = group_cards_by_world(db)
    for world, cards in grouped.items():
        print(f"world={world}: {[card.name for card in cards]}")
