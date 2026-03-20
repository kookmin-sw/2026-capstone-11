from __future__ import annotations

from dataclasses import dataclass, field
from enum import IntEnum
from pathlib import Path
from typing import Any, Dict, List, Optional
import re

import openpyxl

MODULE_DIR = Path(__file__).resolve().parent


class Role(IntEnum):
    LEADER = 0
    BISHOP = 1
    KNIGHT = 2
    ROOK = 3
    PAWN = 4

    @classmethod
    def from_value(cls, value: Any) -> "Role":
        if value is None:
            raise ValueError("Role value is missing.")
        return cls(int(value))


class TextCondition(IntEnum):
    ALWAYS = 0
    TURN_START = 1
    TURN_END = 2
    ON_DESTROYED = 3
    ON_BASIC_MOVE = 4

    @classmethod
    def from_value(cls, value: Any) -> "TextCondition":
        if value is None or value == "":
            return cls.ALWAYS
        return cls(int(value))


ROLE_NAME_KO = {
    Role.LEADER: "군주",
    Role.BISHOP: "비숍",
    Role.KNIGHT: "나이트",
    Role.ROOK: "룩",
    Role.PAWN: "폰",
}

TEXT_CONDITION_NAME_KO = {
    TextCondition.ALWAYS: "상시",
    TextCondition.TURN_START: "턴 시작",
    TextCondition.TURN_END: "턴 종료",
    TextCondition.ON_DESTROYED: "파괴 시",
    TextCondition.ON_BASIC_MOVE: "기본 이동 시",
}


def _normalize_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _normalize_int(value: Any, default: int = 0) -> int:
    if value is None or value == "":
        return default
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, (int, float)):
        return int(value)
    return int(str(value).strip())


def _resolve_module_relative(path: str | Path) -> Path:
    p = Path(path)
    if p.is_absolute():
        return p
    return (MODULE_DIR / p).resolve()


def _contains_any(text: str, keywords: List[str]) -> bool:
    return any(k in text for k in keywords)


def classify_target_schema(text: str) -> str:
    text = _normalize_text(text)
    if not text:
        return "none"
    if "최대 두 마리" in text or "두 마리" in text:
        return "enemy_units_up_to_two"
    if "모든 적" in text or "모든 유닛" in text:
        return "all_units"
    if "같은 행" in text:
        return "same_row"
    if "아군 유닛을 하나 선택" in text or "아군 기물을 하나 선택" in text:
        return "ally_unit"
    if "군주가 아닌 아군" in text:
        return "ally_nonleader_unit"
    if "적 유닛을 하나 선택" in text or "적 기물을 하나 선택" in text:
        return "enemy_unit"
    if "원하는 칸" in text or "위치로 이동" in text or "위치로 옮깁니다" in text:
        return "board_cell"
    if "카드를 1장 뽑" in text and "선택" not in text:
        return "none"
    return "implicit_or_custom"


def extract_effect_flags(text: str) -> Dict[str, int]:
    text = _normalize_text(text)
    flags = {
        "draw": 0,
        "heal": 0,
        "full_heal": 0,
        "move": 0,
        "attack": 0,
        "multi_attack": 0,
        "shield": 0,
        "buff_attack": 0,
        "buff_life": 0,
        "promotion": 0,
        "swap": 0,
        "sacrifice": 0,
        "aoe": 0,
        "disable_move": 0,
        "disable_attack": 0,
        "recall": 0,
        "conditional_same_row": 0,
        "conditional_adjacent_leader": 0,
        "move_after_kill": 0,
    }
    if not text:
        return flags
    if "카드를 1장 뽑" in text or "카드를 한 장 뽑" in text:
        flags["draw"] = 1
    if "체력을 모두 회복" in text:
        flags["full_heal"] = 1
        flags["heal"] = 1
    elif "체력을" in text and "회복" in text:
        flags["heal"] = 1
    if _contains_any(text, ["이동시킵니다", "이동합니다", "위치로 이동", "위치로 옮깁니다"]):
        flags["move"] = 1
    if "공격합니다" in text:
        flags["attack"] = 1
    if _contains_any(text, ["최대 두 마리", "두 마리까지 선택해 공격"]):
        flags["multi_attack"] = 1
    if "보호막" in text:
        flags["shield"] = 1
    if _contains_any(text, ["공격력을", "+", "/"]):
        if "공격력" in text or re.search(r"\+\d+/\+\d+", text):
            flags["buff_attack"] = 1
    if re.search(r"\+\d+/\+\d+", text):
        flags["buff_life"] = 1
    if "프로모션" in text or "체스 퀸의 이동범위" in text:
        flags["promotion"] = 1
    if "위치를 교체" in text:
        flags["swap"] = 1
    if "파괴" in text and ("아군" in text or "희생" in text):
        flags["sacrifice"] = 1
    if "모든 적" in text or "광역" in text:
        flags["aoe"] = 1
    if "이동할 수 없" in text:
        flags["disable_move"] = 1
    if "공격할 수 없" in text:
        flags["disable_attack"] = 1
    if "패로 가져옵니다" in text or "손패로 가져" in text:
        flags["recall"] = 1
    if "같은 행" in text:
        flags["conditional_same_row"] = 1
    if "군주와 인접" in text or "군주와 같은 행" in text:
        flags["conditional_adjacent_leader"] = 1
    if "파괴되면 해당 유닛의 위치로 이동" in text:
        flags["move_after_kill"] = 1
    return flags


@dataclass(frozen=True)
class CardDefinition:
    card_id: str
    name: str
    world: int
    role: Role
    attack: int
    life: int
    text_condition: TextCondition
    text_name: str
    text: str
    effect_name: str
    effect: str
    text_flags: Dict[str, int] = field(default_factory=dict)
    effect_flags: Dict[str, int] = field(default_factory=dict)
    text_target_schema: str = "none"
    effect_target_schema: str = "none"

    @property
    def role_name_ko(self) -> str:
        return ROLE_NAME_KO[self.role]

    @property
    def text_condition_name_ko(self) -> str:
        return TEXT_CONDITION_NAME_KO[self.text_condition]

    def to_dict(self) -> Dict[str, Any]:
        return {
            "card_id": self.card_id,
            "name": self.name,
            "world": self.world,
            "role": int(self.role),
            "role_name_ko": self.role_name_ko,
            "attack": self.attack,
            "life": self.life,
            "text_condition": int(self.text_condition),
            "text_condition_name_ko": self.text_condition_name_ko,
            "text_name": self.text_name,
            "text": self.text,
            "effect_name": self.effect_name,
            "effect": self.effect,
            "text_flags": dict(self.text_flags),
            "effect_flags": dict(self.effect_flags),
            "text_target_schema": self.text_target_schema,
            "effect_target_schema": self.effect_target_schema,
        }

    def to_vector_dict(self) -> Dict[str, Any]:
        return {
            "world": self.world,
            "role": int(self.role),
            "attack": self.attack,
            "life": self.life,
            "text_condition": int(self.text_condition),
            "text_flags": dict(self.text_flags),
            "effect_flags": dict(self.effect_flags),
            "text_target_schema": self.text_target_schema,
            "effect_target_schema": self.effect_target_schema,
        }


def _get_header_map(ws):
    header_row_index = None
    for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if _normalize_text(row[0]) == "CardID":
            header_row_index = idx
            headers = [_normalize_text(c) for c in row]
            break
    if header_row_index is None:
        raise ValueError("Could not find header row starting with 'CardID'.")
    header_map = {}
    for i, h in enumerate(headers):
        if h:
            header_map[h] = i
    return header_map, header_row_index


def load_card_list(xlsx_path: str | Path = "./Cards.xlsx", sheet_name: Optional[str] = None) -> List[CardDefinition]:
    xlsx_path = _resolve_module_relative(xlsx_path)
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]
    header_map, header_row_index = _get_header_map(ws)
    required_headers = ["CardID", "Name", "World", "Role", "Attack", "Life", "TextCondition", "TextName", "Text", "EffectName", "Effect"]
    missing = [h for h in required_headers if h not in header_map]
    if missing:
        raise ValueError(f"Missing required headers: {missing}")

    cards: List[CardDefinition] = []
    for row in ws.iter_rows(min_row=header_row_index + 1, values_only=True):
        card_id = _normalize_text(row[header_map["CardID"]])
        if not card_id:
            continue
        text = _normalize_text(row[header_map["Text"]])
        effect = _normalize_text(row[header_map["Effect"]])
        cards.append(CardDefinition(
            card_id=card_id,
            name=_normalize_text(row[header_map["Name"]]),
            world=_normalize_int(row[header_map["World"]]),
            role=Role.from_value(row[header_map["Role"]]),
            attack=_normalize_int(row[header_map["Attack"]]),
            life=_normalize_int(row[header_map["Life"]]),
            text_condition=TextCondition.from_value(row[header_map["TextCondition"]]),
            text_name=_normalize_text(row[header_map["TextName"]]),
            text=text,
            effect_name=_normalize_text(row[header_map["EffectName"]]),
            effect=effect,
            text_flags=extract_effect_flags(text),
            effect_flags=extract_effect_flags(effect),
            text_target_schema=classify_target_schema(text),
            effect_target_schema=classify_target_schema(effect),
        ))
    return cards


def load_card_db(xlsx_path: str | Path = "./Cards.xlsx", sheet_name: Optional[str] = None) -> Dict[str, CardDefinition]:
    return {card.card_id: card for card in load_card_list(xlsx_path=xlsx_path, sheet_name=sheet_name)}


def group_cards_by_world(card_db: Dict[str, CardDefinition]) -> Dict[int, List[CardDefinition]]:
    grouped: Dict[int, List[CardDefinition]] = {}
    for card in card_db.values():
        grouped.setdefault(card.world, []).append(card)
    for world in grouped:
        grouped[world].sort(key=lambda c: (int(c.role), c.card_id))
    return grouped


def build_default_deck(card_db: Dict[str, CardDefinition], world: int) -> List[str]:
    world_cards = [c for c in card_db.values() if c.world == world]
    if len(world_cards) != 5:
        raise ValueError(f"World {world} does not have exactly 5 prototype cards.")
    by_role = {card.role: card for card in world_cards}
    required_roles = [Role.LEADER, Role.ROOK, Role.KNIGHT, Role.BISHOP, Role.PAWN]
    missing = [r for r in required_roles if r not in by_role]
    if missing:
        missing_names = [ROLE_NAME_KO[r] for r in missing]
        raise ValueError(f"World {world} is missing roles: {missing_names}")
    pawn_id = by_role[Role.PAWN].card_id
    return [
        by_role[Role.LEADER].card_id,
        by_role[Role.ROOK].card_id,
        by_role[Role.KNIGHT].card_id,
        by_role[Role.BISHOP].card_id,
        pawn_id,
        pawn_id,
        pawn_id,
    ]


if __name__ == "__main__":
    db = load_card_db("./Cards.xlsx")
    grouped = group_cards_by_world(db)
    print(f"Loaded {len(db)} cards from {len(grouped)} worlds.")
    for world, cards in grouped.items():
        print(f"\n[World {world}]")
        for card in cards:
            print(
                f"- {card.card_id} | {card.name} | {card.role_name_ko} | "
                f"ATK {card.attack} / LIFE {card.life} | "
                f"TextCond={card.text_condition_name_ko}"
            )
